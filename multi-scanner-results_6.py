import json
import re
import pandas as pd
import argparse
from pathlib import Path
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def clean_url(text):
    """Преобразует текст вида [https://example.com](https://example.com) в https://example.com"""
    # Паттерн для поиска [описание](URL)
    pattern = r'\[([^\]]*)\]\(([^)]+)\)'
    # Заменяем на второй захват (URL)
    return re.sub(pattern, r'\2', text)


def read_sonarqube_excel(file_path):
    """
    Чтение данных из Excel файла SonarQube
    """
    print(f"Чтение файла SonarQube: {file_path}")
    sonarqube_data = []

    try:
        # Читаем лист Issues
        issues_df = pd.read_excel(file_path, sheet_name='Issues')
        print(f"  Прочитано записей с листа Issues: {len(issues_df)}")
        
        # Фильтруем только записи с Type = VULNERABILITY
        if 'Type' in issues_df.columns:
            vulnerability_df = issues_df[issues_df['Type'] == 'VULNERABILITY'].copy()
            print(f"  Найдено уязвимостей (VULNERABILITY): {len(vulnerability_df)}")
            sonarqube_data.append(vulnerability_df)
        else:
            print("  Предупреждение: В файле SonarQube нет колонки 'Type'")
            # Если нет колонки Type, берем все записи
            sonarqube_data.append(issues_df)
            
    except Exception as e:
        print(f"  Ошибка при чтении листа Issues: {e}")

    try:
        # Читаем лист Security Hotspots
        hotspots_df = pd.read_excel(file_path, sheet_name='Security Hotspots')
        print(f"  Прочитано записей с листа Security Hotspots: {len(hotspots_df)}")
        
        # Фильтруем записи: Resolution != 'SAFE'
        if 'Resolution' in hotspots_df.columns:
            # Используем безопасный метод для сравнения
            mask = hotspots_df['Resolution'].astype(str).str.upper() != 'SAFE'
            filtered_hotspots_df = hotspots_df[mask].copy()
            print(f"  Security Hotspots с Resolution != 'SAFE': {len(filtered_hotspots_df)}")
            sonarqube_data.append(filtered_hotspots_df)
        else:
            print("  Предупреждение: В Security Hotspots нет колонки 'Resolution'")
            sonarqube_data.append(hotspots_df)
            
    except Exception as e:
        print(f"  Ошибка при чтении листа Security Hotspots: {e}")

    if not sonarqube_data:
        print("  Предупреждение: Не удалось прочитать данные из файла SonarQube")
        return pd.DataFrame()

    # Объединяем данные с обоих листов
    combined_df = pd.concat(sonarqube_data, ignore_index=True)

    # Выбираем нужные колонки
    required_columns = ['Message', 'File', 'Line']

    # Проверяем, какие колонки существуют
    available_columns = [col for col in required_columns if col in combined_df.columns]

    if not available_columns:
        print("  Предупреждение: В файле SonarQube не найдены требуемые колонки")
        return pd.DataFrame()

    # Создаем DataFrame с нужными колонками
    result_df = combined_df[available_columns].copy()

    # Переименовываем колонки
    rename_dict = {
        'Message': 'Описание',
        'File': 'Файл',
        'Line': 'Строчка'
    }

    # Применяем переименование только для существующих колонок
    rename_dict = {k: v for k, v in rename_dict.items() if k in result_df.columns}
    result_df.rename(columns=rename_dict, inplace=True)

    # Добавляем колонку с источником
    result_df['Источник'] = 'SonarQube'

    # Добавляем пустую колонку для кода (у Sonar Qube нет данных о коде)
    result_df['Код'] = ''

    # Преобразуем значения в колонке "Строчка" в целые числа
    if 'Строчка' in result_df.columns:
        # Сначала преобразуем все значения в строки, затем в числа
        result_df['Строчка'] = pd.to_numeric(result_df['Строчка'], errors='coerce')
        # Заменяем NaN на 0
        result_df['Строчка'] = result_df['Строчка'].fillna(0).astype(int)

    print(f"  Обработано записей SonarQube: {len(result_df)}")
    return result_df


def read_trufflehog_json(file_path):
    """
    Чтение данных из JSON файла Trufflehog
    """
    print(f"Чтение файла Trufflehog: {file_path}")
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            # Читаем строки как отдельные JSON объекты
            lines = f.readlines()
            data_list = []
            
            for line in lines:
                line = line.strip()
                if line:  # Пропускаем пустые строки
                    try:
                        data = json.loads(line)
                        data_list.append(data)
                    except json.JSONDecodeError:
                        continue
        
        print(f"  Прочитано записей Trufflehog: {len(data_list)}")
        
        # Извлекаем нужные данные
        trufflehog_data = []
        for item in data_list:
            # Получаем номер строки
            line_num = item.get('SourceMetadata', {}).get('Data', {}).get('Filesystem', {}).get('line', '')
            try:
                line_num_int = int(line_num) if line_num != '' else 0
            except (ValueError, TypeError):
                line_num_int = 0
            
            # Получаем значение Redacted
            redacted_value = item.get('Redacted', '')
            
            row = {
                'Описание': item.get('DetectorDescription', ''),
                'Строчка': line_num_int,
                'Файл': item.get('SourceMetadata', {}).get('Data', {}).get('Filesystem', {}).get('file', ''),
                'Код': redacted_value
            }
            trufflehog_data.append(row)
        
        # Создаем DataFrame
        if trufflehog_data:
            df = pd.DataFrame(trufflehog_data)
            df['Источник'] = 'Trufflehog'
            print(f"  Обработано записей Trufflehog: {len(df)}")
            return df
        else:
            print("  Предупреждение: В файле Trufflehog не найдено данных")
            return pd.DataFrame()
            
    except Exception as e:
        print(f"  Ошибка при чтении файла Trufflehog: {e}")
        return pd.DataFrame()


def read_gitleaks_json(file_path):
    """
    Чтение данных из JSON файла Gitleaks
    """
    print(f"Чтение файла Gitleaks: {file_path}")
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data_list = json.load(f)
        
        print(f"  Прочитано записей Gitleaks: {len(data_list)}")
        
        # Извлекаем нужные данные
        gitleaks_data = []
        for item in data_list:
            # Получаем номер строки
            line_num = item.get('StartLine', '')
            try:
                line_num_int = int(line_num) if line_num != '' else 0
            except (ValueError, TypeError):
                line_num_int = 0
            
            # Получаем значение Match
            match_value = item.get('Match', '')
            
            row = {
                'Описание': item.get('Description', ''),
                'Строчка': line_num_int,
                'Файл': item.get('File', ''),
                'Код': match_value
            }
            gitleaks_data.append(row)
        
        # Создаем DataFrame
        if gitleaks_data:
            df = pd.DataFrame(gitleaks_data)
            df['Источник'] = 'Gitleaks'
            print(f"  Обработано записей Gitleaks: {len(df)}")
            return df
        else:
            print("  Предупреждение: В файле Gitleaks не найдено данных")
            return pd.DataFrame()
            
    except Exception as e:
        print(f"  Ошибка при чтении файла Gitleaks: {e}")
        return pd.DataFrame()


def read_semgrep_sarif(file_path):
    """
    Чтение данных из SARIF файла Semgrep
    """
    print(f"Чтение файла Semgrep: {file_path}")
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        semgrep_data = []
        
        # Проходим по всем runs в SARIF отчете
        runs = data.get('runs', [])
        print(f"  Найдено runs: {len(runs)}")
        
        for run in runs:
            # Получаем результаты из текущего run
            results = run.get('results', [])
            
            # Создаем словарь для быстрого доступа к правилам
            rules_dict = {}
            rules = run.get('tool', {}).get('driver', {}).get('rules', [])
            for rule in rules:
                rule_id = rule.get('id')
                if rule_id:
                    rules_dict[rule_id] = rule
            
            for result in results:
                # Извлекаем описание из help.markdown
                description = ''
                rule_id = result.get('ruleId', '')
                priority_level = ''  # Добавляем переменную для уровня приоритета
                
                # Получаем правило из предварительно созданного словаря
                if rule_id in rules_dict:
                    rule = rules_dict[rule_id]
                    help_markdown = rule.get('help', {}).get('markdown', '')
                    if help_markdown:
                        # Берем полное описание без обрезки
                        description = help_markdown.strip()
                    else:
                        # Если нет help.markdown, берем shortDescription или fullDescription
                        description = rule.get('shortDescription', {}).get('text', '')
                        if not description:
                            description = rule.get('fullDescription', {}).get('text', '')
                    
                    # Извлекаем уровень приоритета (level) из properties или напрямую из rule
                    # Уровень может быть в разных местах в зависимости от версии SARIF
                    properties = rule.get('properties', {})
                    
                    # Пробуем получить level из разных возможных мест
                    if 'level' in properties:
                        priority_level = properties.get('level', '')
                    elif 'severity' in properties:
                        priority_level = properties.get('severity', '')
                    elif 'defaultConfiguration' in rule:
                        priority_level = rule.get('defaultConfiguration', {}).get('level', '')
                    
                    # Если уровень найден, преобразуем для единообразия
                    if priority_level:
                        # Приводим к верхнему регистру для единообразия
                        priority_level = priority_level.upper()
                
                # Если описание не найдено, используем ruleId
                if not description:
                    description = f"Semgrep rule: {rule_id}"
                
                # Если уровень не найден, пробуем получить из result
                if not priority_level:
                    priority_level = result.get('level', '').upper()
                
                # Извлекаем информацию о location и snippet
                locations = result.get('locations', [])
                for location in locations:
                    physical_location = location.get('physicalLocation', {})
                    artifact_location = physical_location.get('artifactLocation', {})
                    file_path_sarif = artifact_location.get('uri', '')
                    
                    # Получаем номер строки
                    region = physical_location.get('region', {})
                    start_line = region.get('startLine', 0)
                    
                    try:
                        start_line_int = int(start_line) if start_line else 0
                    except (ValueError, TypeError):
                        start_line_int = 0
                    
                    # Получаем snippet (фрагмент кода)
                    snippet = result.get('codeFlows', [])
                    code_snippet = ''
                    
                    # Пробуем получить snippet из разных мест в SARIF
                    if snippet and len(snippet) > 0:
                        # Извлекаем из codeFlows
                        threads = snippet[0].get('threadFlows', [])
                        if threads and len(threads) > 0:
                            locations_list = threads[0].get('locations', [])
                            if locations_list and len(locations_list) > 0:
                                code_snippet = locations_list[0].get('location', {}).get('physicalLocation', {}).get('region', {}).get('snippet', {}).get('text', '')
                     
                    # Если не нашли в codeFlows, пробуем получить напрямую из result
                    if not code_snippet:
                        code_snippet = result.get('snippet', {}).get('text', '')
                    
                    # Если всё еще нет, пробуем из location region snippet
                    if not code_snippet:
                        code_snippet = region.get('snippet', {}).get('text', '')
                    
                    row = {
                        'Описание': description,
                        'Строчка': start_line_int,
                        'Файл': file_path_sarif,
                        'Код': code_snippet,
                        'Приоритет': priority_level  # Добавляем уровень приоритета
                    }
                    semgrep_data.append(row)
        
        # Создаем DataFrame
        if semgrep_data:
            df = pd.DataFrame(semgrep_data)
            df['Источник'] = 'Semgrep'
            # Удаляем возможные дубликаты внутри Semgrep
            df = df.drop_duplicates(subset=['Описание', 'Файл', 'Строчка'], keep='first')
            print(f"  Обработано записей Semgrep: {len(df)}")
            # Выводим информацию о найденных уровнях приоритета
            if 'Приоритет' in df.columns and not df['Приоритет'].empty:
                unique_levels = df['Приоритет'].unique()
                print(f"  Найдены уровни приоритета: {', '.join(filter(None, unique_levels))}")
            return df
        else:
            print("  Предупреждение: В файле Semgrep не найдено данных")
            return pd.DataFrame()
            
    except Exception as e:
        print(f"  Ошибка при чтении файла Semgrep: {e}")
        return pd.DataFrame()


def combine_all_data(sonarqube_df, trufflehog_df, gitleaks_df, semgrep_df):
    """
    Объединение данных из всех источников с удалением дублей
    """
    print("\nОбъединение данных из всех источников...")
    # Создаем список всех DataFrames
    all_dataframes = []

    if not sonarqube_df.empty:
        all_dataframes.append(sonarqube_df)
        print(f"  Добавлены данные SonarQube: {len(sonarqube_df)} записей")

    if not trufflehog_df.empty:
        all_dataframes.append(trufflehog_df)
        print(f"  Добавлены данные Trufflehog: {len(trufflehog_df)} записей")

    if not gitleaks_df.empty:
        all_dataframes.append(gitleaks_df)
        print(f"  Добавлены данные Gitleaks: {len(gitleaks_df)} записей")

    if not semgrep_df.empty:
        all_dataframes.append(semgrep_df)
        print(f"  Добавлены данные Semgrep: {len(semgrep_df)} записей")

    if not all_dataframes:
        print("  Внимание: Нет данных для объединения!")
        return pd.DataFrame()

    # Объединяем все DataFrames
    combined_df = pd.concat(all_dataframes, ignore_index=True)

    # Удаляем дубли на основе комбинации Описание, Файл и Строчка
    print(f"\n  Записей до удаления дублей: {len(combined_df)}")

    # Создаем временную колонку для сравнения
    combined_df['_temp_key'] = combined_df['Описание'].fillna('').astype(str) + '|' + \
                               combined_df['Файл'].fillna('').astype(str) + '|' + \
                               combined_df['Строчка'].astype(str)

    # Удаляем дубли
    combined_df = combined_df.drop_duplicates(subset='_temp_key', keep='first')

    # Удаляем временную колонку
    combined_df = combined_df.drop(columns=['_temp_key'])

    print(f"  Записей после удаления дублей: {len(combined_df)}")

    # Если в объединенных данных нет колонки "Приоритет", создаем её
    if 'Приоритет' not in combined_df.columns:
        combined_df['Приоритет'] = ''

    # Определяем порядок колонок (Код после Файл, Приоритет перед Примечание)
    column_order = ['Источник', 'Описание', 'Файл', 'Код', 'Строчка', 'Статус', 'Приоритет', 'Примечание']

    # Создаем финальный DataFrame с нужным порядком колонок
    final_df = pd.DataFrame()

    for col in column_order:
        if col in combined_df.columns:
            final_df[col] = combined_df[col]
        else:
            # Если колонки нет, создаем пустую
            final_df[col] = ''

    # Преобразуем значения в колонке "Строчка" в целые числа (на всякий случай)
    if 'Строчка' in final_df.columns:
        final_df['Строчка'] = pd.to_numeric(final_df['Строчка'], errors='coerce').fillna(0).astype(int)

    print(f"\nИтого объединено записей: {len(final_df)}")
    return final_df


def apply_translation_to_dataframe(df, translations_map):
    """
    Применяет перевод к DataFrame, заменяя значения в колонке 'Описание'.
    """
    print("Применение перевода к DataFrame...")
    new_entries_to_add = []
    for idx, row in df.iterrows():
        original_text = str(row['Описание']) # Преобразуем в строку на всякий случай
        clean_original_text = clean_url(original_text.strip())

        # Проверить, есть ли перевод в словаре
        if clean_original_text in translations_map:
            # Получаем русский перевод
            ru_translation = translations_map[clean_original_text]
            if ru_translation: # Если 'ru' не пустой
                # Заменить значение в DataFrame на русский перевод
                df.at[idx, 'Описание'] = ru_translation
            # else: если 'ru' пустой, оставляем английский текст как есть
        else:
            # Если перевод не найден, добавить в список новых
            # только если такой ключ ещё не был добавлен ранее в этот цикл
            if clean_original_text not in [entry['en'] for entry in new_entries_to_add]:
                # Добавляем новую запись в список для добавления в JSON с пустым 'ru'
                new_entry = {"en": original_text, "ru": ""}
                new_entries_to_add.append(new_entry)
    print(f"Обновлено {len(df) - sum(1 for x in df['Описание'] if x in [entry['en'] for entry in new_entries_to_add])} записей с переводом.")
    print(f"Найдено {len(new_entries_to_add)} новых записей для словаря переводов.")
    return df, new_entries_to_add


def update_translations_json_and_save(new_entries_to_add, translations_list, existing_translations_map, json_filename='translations.json'):
    """
    Обновляет и сохраняет translations.json, добавляя новые записи в конец.
    """
    print("Обновление файла translations.json...")
    # Обновить список translations_list, добавив новые записи в конец
    # Не добавляем дубликаты по ключу 'en' (очищенному), которые уже есть в JSON.
    unique_new_entries = []
    seen_new_keys = set()
    for entry in new_entries_to_add:
        clean_key = clean_url(entry['en'].strip())
        if clean_key not in existing_translations_map and clean_key not in seen_new_keys:
            unique_new_entries.append(entry)
            seen_new_keys.add(clean_key)

    # Объединяем старый список с уникальными новыми записями
    updated_translations_list = translations_list + unique_new_entries

    # Обновить и сохранить translations.json
    data_for_saving = {'translations': updated_translations_list}
    # Сохраняем с отступом для лучшего оформления, не меняя порядок ключей
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(data_for_saving, f, ensure_ascii=False, indent=2)
    print(f"Файл {json_filename} обновлён. Всего записей: {len(updated_translations_list)}. Новых: {len(unique_new_entries)}.")
    return True


def format_excel_with_styles(file_path):
    """
    Форматирование Excel файла: стили заголовков, границы, ширина колонок, фильтры
    """
    try:
        # Загружаем workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Определяем цвета
        # Оранжевый 80% оттенка (RGB: 255, 200, 100 примерно 80% от яркости)
        orange_fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
        
        # Определяем стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Определяем стиль для заголовков (жирный шрифт)
        header_font = Font(bold=True)
        
        # Определяем выравнивание по центру для заголовков (горизонтальное и вертикальное)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Определяем выравнивание по центру для данных (вертикальное по центру)
        data_center_vertical_alignment = Alignment(vertical='center', wrap_text=True)
        
        # Определяем выравнивание по центру для колонки "Строчка" (горизонтальное и вертикальное)
        line_number_alignment = Alignment(horizontal='center', vertical='center')
        
        # Применяем стили к заголовкам (первая строка)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Применяем границы и выравнивание ко всем ячейкам с данными
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                # Получаем имя колонки
                col_name = ws.cell(row=1, column=col).value
                
                # Для колонки "Строчка" устанавливаем выравнивание по центру
                if col_name in ['Строчка', 'Приоритет']:
                    cell.alignment = line_number_alignment
                else:
                    # Для остальных колонок - вертикальное выравнивание по центру
                    # Если это колонки с длинным текстом, включаем перенос
                    if col_name in ['Описание', 'Код', 'Примечание', 'Файл']:
                        cell.alignment = data_center_vertical_alignment
                    else:
                        # Для коротких колонок - просто вертикальное выравнивание по центру
                        cell.alignment = Alignment(vertical='center')
        
        # Настраиваем ширину колонок
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            col_name = ws.cell(row=1, column=col).value
            
            # Находим колонку "Источник" для автоширины
            if col_name == 'Источник':
                # Автоширина для колонки Источник
                max_length = 0
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                # Добавляем небольшой отступ
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Для остальных колонок устанавливаем стандартную ширину
            else:
                if col_name == 'Описание':
                    ws.column_dimensions[col_letter].width = 50
                elif col_name == 'Файл':
                    ws.column_dimensions[col_letter].width = 40
                elif col_name == 'Код':
                    ws.column_dimensions[col_letter].width = 60
                elif col_name == 'Строчка':
                    ws.column_dimensions[col_letter].width = 10
                elif col_name == 'Статус':
                    ws.column_dimensions[col_letter].width = 12
                elif col_name == 'Приоритет':
                    ws.column_dimensions[col_letter].width = 12
                elif col_name == 'Примечание':
                    ws.column_dimensions[col_letter].width = 30
                else:
                    ws.column_dimensions[col_letter].width = 15
        
        # Добавляем автофильтр для всех колонок
        # Определяем диапазон: от A1 до последней колонки и последней строки
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row > 1:  # Если есть данные кроме заголовка
            filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = filter_range
            print(f"  Добавлен автофильтр для диапазона: {filter_range}")
        
        # Дополнительно: устанавливаем высоту строк для лучшего отображения
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = None  # Автовысота
         
        # Сохраняем изменения
        wb.save(file_path)
        print("  Применено форматирование: границы, цвета заголовков, ширина колонок, перенос текста, вертикальное выравнивание по центру, автофильтр")
        return True
        
    except Exception as e:
        print(f"  Ошибка при форматировании Excel: {e}")
        return False


def save_to_excel(dataframe, output_file):
    """
    Сохранение данных в Excel файл с последующим форматированием
    """
    if dataframe.empty:
        print("Нет данных для сохранения!")
        return False
    try:
        # Сохраняем в Excel
        dataframe.to_excel(output_file, index=False)
        print(f"\nДанные успешно сохранены в файл: {output_file}")
        print(f"Количество записей: {len(dataframe)}")
        
        # Выводим статистику по источникам
        if 'Источник' in dataframe.columns:
            print("\nСтатистика по источникам:")
            source_stats = dataframe['Источник'].value_counts()
            for source, count in source_stats.items():
                print(f"  {source}: {count} записей")
        
        return True
        
    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {e}")
        return False


def load_translations(json_filename='translations.json'):
    """
    Загружает translations.json и возвращает словарь {cleaned_en: ru}.
    """
    print(f"Загрузка файла переводов: {json_filename}")
    try:
        with open(json_filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Файл {json_filename} не найден. Создаю пустой словарь.")
        return {}, []
    except json.JSONDecodeError as e:
        print(f"Ошибка чтения JSON из {json_filename}: {e}")
        return {}, []

    translations_list = data.get('translations', [])
    # Создать словарь {очищенный_ключ: ru_translation} для существующих переводов
    existing_translations_map = {}
    for item in translations_list:
        en_raw = item.get('en', '').strip()
        en_clean = clean_url(en_raw)
        ru_translation = item.get('ru', '')
        # Если в JSON уже есть запись с пустым 'ru', она все равно добавляется в карту.
        # При поиске в цикле apply_translation_to_dataframe будет найдена, и если в 'ru' пусто,
        # текст в Excel останется без изменений, и новая запись в список для JSON добавлена не будет.
        existing_translations_map[en_clean] = ru_translation

    print(f"Загружено {len(existing_translations_map)} переводов.")
    return existing_translations_map, translations_list


def main():
    # Настройка аргументов командной строки
    parser = argparse.ArgumentParser(description='Объединение данных из различных источников безопасности с переводом')
    parser.add_argument('--sonarqube', '-s', help='Путь к файлу SonarQube Excel', default='2026-02-16-VTZ.MES-issues-report.xlsx')
    parser.add_argument('--trufflehog', '-t', help='Путь к файлу Trufflehog JSON', default='trufflehog_secret_scan.json')
    parser.add_argument('--gitleaks', '-g', help='Путь к файлу Gitleaks JSON', default='gitleaks-dir-report.json')
    parser.add_argument('--semgrep', '-m', help='Путь к файлу Semgrep SARIF', default='semgrep-report.sarif')
    parser.add_argument('--output', '-o', help='Путь к выходному файлу', default='results.xlsx')
    parser.add_argument('--translations-json', '--tj', help='Путь к файлу translations.json', default='translations.json')

    args = parser.parse_args()

    print("=" * 60)
    print("Сбор данных из различных источников безопасности с переводом")
    print("=" * 60)
    print("Примечания:")
    print("- Из SonarQube Issues берутся только записи с Type = VULNERABILITY")
    print("- Из SonarQube Security Hotspots берутся записи с Resolution != 'SAFE'")
    print("- Все значения в колонке 'Строчка' преобразуются в целые числа")
    print("- Дубли удаляются после объединения всех данных")
    print("- Semgrep: из SARIF файла извлекаются description, file, startLine и snippet")
    print("- Trufflehog: из JSON извлекается Redacted")
    print("- Gitleaks: из JSON извлекается Match")
    print("- Выходной Excel файл форматируется: границы, заголовок оранжевый, автоширина")
    print("- Перевод выполняется перед форматированием Excel")
    print("=" * 60)

    # Чтение данных из всех источников
    sonarqube_df = read_sonarqube_excel(args.sonarqube)
    trufflehog_df = read_trufflehog_json(args.trufflehog)
    gitleaks_df = read_gitleaks_json(args.gitleaks)
    semgrep_df = read_semgrep_sarif(args.semgrep)

    # Объединение данных с удалением дублей
    combined_df = combine_all_data(sonarqube_df, trufflehog_df, gitleaks_df, semgrep_df)

    if combined_df.empty:
        print("Нет данных для обработки.")
        return

    # Загрузка словаря переводов
    existing_translations_map, translations_list = load_translations(args.translations_json)

    # Применение перевода к DataFrame
    translated_df, new_entries_to_add = apply_translation_to_dataframe(combined_df, existing_translations_map)

    # Обновление и сохранение translations.json
    update_translations_json_and_save(new_entries_to_add, translations_list, existing_translations_map, args.translations_json)

    # Сохранение результатов в Excel
    success = save_to_excel(translated_df, args.output)

    if success:
        print("\nПрименение форматирования к Excel файлу...")
        format_excel_with_styles(args.output)

    print("\n" + "=" * 60)
    print("Завершено!")
    print("=" * 60)


if __name__ == "__main__":
    main()